movies = ['电影1', '电影2', '电影3', '电影4', '电影5', '电影6', '电影7', '电影8', '电影9', '电影10',
          '电影11', '电影12', '电影13', '电影14', '电影15', '电影16', '电影17', '电影18', '电影19', '电影20']
director = ['导演1', '导演2', '导演3', '导演4', '导演5', '导演6', '导演7', '导演8', '导演9', '导演10',
            '导演11', '导演12', '导演13', '导演14', '导演15', '导演16', '导演17', '导演18', '导演19', '导演20']
actor = ['演员1', '演员2', '演员3', '演员4', '演员5', '演员6', '演员7', '演员8',
         '演员9', '演员10', '演员11', '演员12', '演员13', '演员14', '演员15']
import random
from openpyxl import Workbook, load_workbook


# 实例化
wb = Workbook()
# 激活 worksheet
ws = wb.active
# 设置表头
ws.append(['电影名称', '导演', '演员'])
col = ws.column_dimensions['C']  # 将演员列拓宽
col.width = 140
for i in range(20):
    acs = []
    ac = []  # 每次执行时均清空
    num = random.randint(4, 15)  # 生成演员数量
    while len(acs) < num:
        x = random.randint(0, 14)
        if x not in acs:  # 去重
            acs.append(x)
            ac.append(actor[x])
            # 演员输出格式化
            ac2 = ','.join(str(i) for i in ac)
    # 生成第i行数据
    ws.append([movies[i], director[i], ac2])

wb.save('电影导演演员信息.xlsx')


actor_dict = dict()
# 遍历Excel文件中的所有行
for index, row in enumerate(ws.rows):
    # 跳过表头，对于每一行有效数据，获取每一行的电影名称和演员清单，
    if index == 0:
        continue
    # 获取电影名称和演员列表
    filmName, actor = row[0].value, row[2].value.split(',')
    # 遍历该电影的所有演员，统计参演电影
    for a in actor:
        actor_dict[a] = actor_dict.get(a, set())
        actor_dict[a].add(filmName)
actor_list = sorted(actor_dict.items(), key=lambda x: int(x[0][2:]))
actor_dict = dict(actor_list)

print("演员所参演的电影字典：{}".format(actor_dict))


actorPairs = []
# 遍历演员和参演电影
for index, actor1 in enumerate(actor_list[:-1]):
    for _, actor2 in enumerate(actor_list[index + 1:]):
        # 构建演员组合演员组合，这里方便起见设置n=2
        actorPair = (actor1[0], actor2[0])
        # 共同参演的电影
        films = actor1[1] & actor2[1]
        # 将 演员组合:电影集合 键值对填入序列
        actorPairs.append((actorPair, films))

# 找到一同参演电影次数最多的组合
favoriPair = max(actorPairs, key=lambda item: len(item[1]))
print("关系中最好的2个演员{},共同参演电影数量{}".format(favoriPair[0], len(favoriPair[1])))



# 打开文件及表
wb = load_workbook('电影导演演员信息.xlsx')
ws = wb["Sheet"]

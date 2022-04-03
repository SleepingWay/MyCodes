from openpyxl import Workbook, load_workbook
import random

movies = ['电影1', '电影2', '电影3', '电影4', '电影5', '电影6', '电影7', '电影8', '电影9', '电影10',
          '电影11', '电影12', '电影13', '电影14', '电影15', '电影16', '电影17', '电影18', '电影19', '电影20', ]
users = ['用户1', '用户2', '用户3', '用户4', '用户5', '用户6', '用户7', '用户8', '用户9', '用户10', ]


def DataCreate():
    wb = Workbook()
    ws = wb.active
    ws.append(['用户', '电影', '打分'])
    userNum = random.randint(5, 10)  # 用户数目
    movieNum = random.randint(3, 20)  # 用户打分的电影数目
    for i in range(userNum):
        user = random.choice(users)
        movie = set(random.choices(movies, k=movieNum))
        for x in movie:
            ws.append([user, x, random.randint(1, 5)])
    wb.save('用户打分信息.xlsx')


def FindFamiliar():
    wb = load_workbook('用户打分信息.xlsx')
    ws = wb["Sheet"]
    user = []  # 读取用户
    movieDict = {}
    row = ws.rows
    for x in row:
        if x[0].value not in user and x[0].value != '用户':
            user.append(x[0].value)
    for x in user:  # 为每个用户建立电影字典
        if x not in movieDict.keys():
            movieDict[x] = []
        row = ws.rows
        for y in row:
            if y[0].value == x:
                movieDict[x].append(y[1].value)
    user1, user2 = 0, 0  # 最大相同数时的两个用户
    maxNum = 0  # 最大相同数
    score1, score2 = 0, 0  # 两个用户的打分和
    minScore = 999999  # 最小的打分和之差
    lst = []
    for x in movieDict.keys():
        nowNum = 0
        for y in movieDict.keys():
            if y != x:
                lst = [m for m in movieDict[x] if m in movieDict[x] and m in movieDict[y]]
                nowNum = len(lst)
                if nowNum > maxNum:
                    maxNum = nowNum
                    user1 = x
                    user2 = y
                    row = ws.rows
                    for z in row:
                        if z[0].value == x and z[1].value in lst:
                            score1 += z[2].value
                        if z[0].value == y and z[1].value in lst:
                            score2 += z[2].value
                    minScore = abs(score1 - score2)  # 计算打分和之差留做备用
                if nowNum == maxNum:  # 如果共同打分的电影数相同，则看打分和之差，越小的则越相似
                    row = ws.rows
                    for z in row:
                        if z[0].value == x and z[1].value in lst:
                            score1 += z[2].value
                        if z[0].value == y and z[1].value in lst:
                            score2 += z[2].value
                    if abs(score1 - score2) < minScore:
                        minScore = abs(score1 - score2)
                        user1 = x
                        user2 = y
    row = ws.rows
    recommendMovie = 0  # 推荐的电影
    maxScore = 0    # 打的最高分
    if len(movieDict[user1]) < len(movieDict[user2]):  # user1最终为为打分电影最多的
        user1, user2 = user2, user1
    for x in row:
        if x[0].value == user1 and x[1].value not in lst:
            if x[2].value >= maxScore:
                recommendMovie = x[1].value
                maxScore = x[2].value
    print('通过筛查，最相似的用户为：{} 和 {}'.format(user1, user2))
    print('{} 推荐的电影为 {} '.format(user1, recommendMovie))


if __name__ == '__main__':
    DataCreate()
    FindFamiliar()
